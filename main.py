import os
import sys
import argparse
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ==============================
# Excel Automation Tool
# ==============================

APP_NAME = "Excel Automation Tool"


# ---------- 选择 Excel ----------
def choose_files():
    files = filedialog.askopenfilenames(
        title="选择 Excel 文件",
        filetypes=[
            ("Excel Files", "*.xlsx *.xls"),
            ("All Files", "*.*")
        ]
    )

    if files:
        file_list.delete(0, tk.END)

        for file in files:
            file_list.insert(tk.END, file)


# ---------- 读取全部 Excel ----------
def read_all_excel(files):

    all_data = []

    for file in files:

        try:
            df = pd.read_excel(file)

            # 加入来源文件
            df["Source File"] = os.path.basename(file)

            all_data.append(df)

        except Exception as e:
            print(f"读取失败: {file}")
            print(e)

    if not all_data:
        raise Exception("没有成功读取任何 Excel 文件。")

    result = pd.concat(
        all_data,
        ignore_index=True
    )

    return result


# ---------- 清理资料 ----------
def detect_and_format_dates(df):
    """
    在 main.py 中复用的日期检测/格式化：支持 Excel 序列和字符串解析，保留无法解析的原值。
    """
    def _parse_cell(val):
        if pd.isna(val):
            return pd.NaT
        if isinstance(val, (int, float)):
            try:
                return pd.to_datetime(val, unit='d', origin='1899-12-30')
            except Exception:
                pass
        try:
            return pd.to_datetime(val, errors='coerce')
        except Exception:
            return pd.NaT

    for col in df.columns:
        try:
            parsed = df[col].apply(_parse_cell)
        except Exception:
            continue

        non_null_ratio = parsed.notna().sum() / max(1, len(parsed))
        if non_null_ratio >= 0.4:
            try:
                formatted = parsed.dt.strftime('%Y-%m-%d')
                df[col] = formatted.where(parsed.notna(), df[col].astype(str))
            except Exception:
                df[col] = [d.strftime('%Y-%m-%d') if not pd.isna(d) else str(orig) for d, orig in zip(parsed, df[col])]

    return df


def clean_data(df):

    # 删除完全空白行
    df = df.dropna(how="all")

    # 清理栏位名称
    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    # 清理文字
    for column in df.columns:

        if df[column].dtype == "object":

            df[column] = (
                df[column]
                .astype(str)
                .str.strip()
            )

    # 标准化日期列（新加）
    df = detect_and_format_dates(df)

    # 删除完全重复资料
    df = df.drop_duplicates()

    return df


# ---------- 自动寻找 ID ----------
def find_id_column(df):

    possible_names = [
        "ID",
        "Id",
        "id",
        "编号",
        "编号ID",
        "Code",
        "code",
        "No",
        "NO",
        "Number",
        "号码"
    ]

    for name in possible_names:

        if name in df.columns:
            return name

    return None


# ---------- 自动排序 ----------
def sort_data(df):

    id_column = find_id_column(df)

    # 如果有日期，优先按日期
    date_columns = [
        column
        for column in df.columns
        if "date" in column.lower()
        or "日期" in column
    ]

    if date_columns:

        column = date_columns[0]

        try:
            df[column] = pd.to_datetime(
                df[column],
                errors="coerce"
            )

            df = df.sort_values(
                by=column,
                na_position="last"
            )

        except Exception:
            pass

    # 如果有 ID，再按 ID 排序
    if id_column:

        df = df.sort_values(
            by=id_column,
            na_position="last"
        )

    return df


# ---------- 找重复 ----------
def find_duplicates(df):

    id_column = find_id_column(df)

    if id_column:

        duplicates = df[
            df[id_column].duplicated(
                keep=False
            )
        ]

    else:

        duplicates = df[
            df.duplicated(
                keep=False
            )
        ]

    return duplicates


# ---------- 自动统计 ----------
def create_summary(
    original_count,
    final_count,
    duplicate_count,
    column_count
):

    summary = pd.DataFrame({

        "项目": [
            "原始资料数量",
            "整理后数量",
            "删除重复数量",
            "栏位数量"
        ],

        "数量": [
            original_count,
            final_count,
            duplicate_count,
            column_count
        ]
    })

    return summary


# ---------- Excel 美化 ----------
def format_excel(output_file):

    wb = load_workbook(output_file)

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    blue = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    # ---------- Data ----------
    if "Data" in wb.sheetnames:

        ws = wb["Data"]

        # 标题
        for cell in ws[1]:

            cell.fill = blue
            cell.font = white_font

        # 自动宽度
        for column in ws.columns:

            max_length = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    length = len(
                        str(cell.value)
                    )

                    if length > max_length:
                        max_length = length

                except Exception:
                    pass

            ws.column_dimensions[
                letter
            ].width = min(
                max_length + 2,
                40
            )

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

    # ---------- Duplicates ----------
    if "Duplicates" in wb.sheetnames:

        ws = wb["Duplicates"]

        for cell in ws[1]:

            cell.fill = red
            cell.font = white_font

        for column in ws.columns:

            letter = get_column_letter(
                column[0].column
            )

            ws.column_dimensions[
                letter
            ].width = 20

    # ---------- Summary ----------
    if "Summary" in wb.sheetnames:

        ws = wb["Summary"]

        for cell in ws[1]:

            cell.fill = blue
            cell.font = white_font

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    wb.save(output_file)


# ---------- 开始自动化 ----------
def start_automation():

    try:

        files = file_list.get(0, tk.END)

        if not files:

            messagebox.showwarning(
                "提醒",
                "请先选择 Excel 文件。"
            )

            return

        status_label.config(
            text="正在读取 Excel...",
            fg="blue"
        )

        root.update()

        # 读取
        df = read_all_excel(files)

        original_count = len(df)

        # 清理
        df = clean_data(df)

        # 找重复
        duplicates = find_duplicates(df)

        duplicate_count = len(duplicates)

        # 删除重复
        df = df.drop_duplicates()

        # 排序
        df = sort_data(df)

        final_count = len(df)

        column_count = len(df.columns)

        # Summary
        summary = create_summary(
            original_count,
            final_count,
            duplicate_count,
            column_count
        )

        # ---------- Output ----------
        first_file = files[0]

        base_folder = os.path.dirname(
            first_file
        )

        output_folder = os.path.join(
            base_folder,
            "Output"
        )

        os.makedirs(
            output_folder,
            exist_ok=True
        )

        output_file = os.path.join(
            output_folder,
            "Excel_Automation_Result.xlsx"
        )

        # ---------- 写入 Excel ----------
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Data",
                index=False
            )

            summary.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            if not duplicates.empty:

                duplicates.to_excel(
                    writer,
                    sheet_name="Duplicates",
                    index=False
                )

        # 美化
        format_excel(output_file)

        status_label.config(
            text="处理完成！",
            fg="green"
        )

        messagebox.showinfo(
            "完成",
            f"Excel 自动化完成！\n\n"
            f"原始资料：{original_count}\n"
            f"整理后：{final_count}\n"
            f"重复资料：{duplicate_count}\n\n"
            f"结果：\n{output_file}"
        )

    except Exception as error:

        status_label.config(
            text="发生错误",
            fg="red"
        )

        messagebox.showerror(
            "错误",
            str(error)
        )


# ==============================
# CLI Helper
# ==============================

def cli_main(argv=None):
    parser = argparse.ArgumentParser(description='Excel Automation Tool - CLI')
    parser.add_argument('-i', '--input', help='单个 Excel 文件或目录（配合 --batch）')
    parser.add_argument('-o', '--output-dir', help='输出目录，默认为第一输入文件所在目录')
    parser.add_argument('--batch', action='store_true', help='如果输入是目录，则批量处理目录下所有 .xlsx/.xls 文件')

    args = parser.parse_args(argv)

    if args.input:
        paths = []
        if os.path.isdir(args.input):
            if args.batch:
                for fn in os.listdir(args.input):
                    if fn.lower().endswith(('.xlsx', '.xls')):
                        paths.append(os.path.join(args.input, fn))
            else:
                print('输入是目录，请加 --batch 以批量处理，或指定单个文件路径')
                sys.exit(1)
        elif os.path.isfile(args.input):
            paths = [args.input]
        else:
            print('输入路径无效')
            sys.exit(1)

        for p in paths:
            try:
                df = read_all_excel([p])
                original_count = len(df)
                df = clean_data(df)
                duplicates = find_duplicates(df)
                duplicate_count = len(duplicates)
                df = df.drop_duplicates()
                df = sort_data(df)
                final_count = len(df)

                base_folder = os.path.dirname(p)
                output_folder = args.output_dir or os.path.join(base_folder, 'Output')
                os.makedirs(output_folder, exist_ok=True)
                output_file = os.path.join(output_folder, 'Excel_Automation_Result.xlsx')

                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Data', index=False)
                    create_summary(original_count, final_count, duplicate_count, len(df.columns)).to_excel(writer, sheet_name='Summary', index=False)
                    if not duplicates.empty:
                        duplicates.to_excel(writer, sheet_name='Duplicates', index=False)

                format_excel(output_file)
                print(f'处理完成：{p} -> {output_file}')
            except Exception as e:
                print(f'处理失败: {p} -> {e}')
        return


# ==============================
# GUI
# ==============================

root = tk.Tk()

root.title(APP_NAME)

root.geometry(
    "750x550"
)

root.resizable(
    False,
    False
)


# 标题
title = tk.Label(
    root,
    text="Excel Automation Tool",
    font=("Arial", 24, "bold")
)

title.pack(pady=25)


subtitle = tk.Label(
    root,
    text="自动读取、整理、排序、去重复、产生报告",
    font=("Arial", 12)
)

subtitle.pack()


# 文件列表
file_list = tk.Listbox(
    root,
    width=85,
    height=12
)

file_list.pack(
    padx=30,
    pady=20
)


# 选择按钮
choose_button = tk.Button(
    root,
    text="选择 Excel 文件",
    command=choose_files,
    font=("Arial", 12),
    padx=20,
    pady=8
)

choose_button.pack(
    pady=5
)


# 开始按钮
start_button = tk.Button(
    root,
    text="开始自动处理",
    command=start_automation,
    font=("Arial", 15, "bold"),
    bg="#4472C4",
    fg="white",
    padx=35,
    pady=12
)

start_button.pack(
    pady=20
)


# 状态
status_label = tk.Label(
    root,
    text="等待选择 Excel...",
    font=("Arial", 11),
    fg="gray"
)

status_label.pack()


# 启动
if len(sys.argv) > 1:
    cli_main()
else:
    root.mainloop()
